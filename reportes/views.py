# reportes/views.py

import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta
from math import sqrt
from django.views.generic import TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
# SE AGREGÓ 'Q' A LA IMPORTACIÓN
from django.db.models import Count, Avg, StdDev, Q
from django.db.models.functions import TruncDate
from django.template.loader import get_template
from xhtml2pdf import pisa
from clinica.models import Paciente, Parto, RecienNacido, TipoParto
from clinica.models import HistorialPaciente

# --- IMPORTACIÓN SEGURIDAD ---
from core.mixins import PermitsPositionMixin

# --- MIXIN DE FILTRADO ---
class ReporteFilterMixin:
    def get_filtered_querysets(self, request):
        fecha_inicio_str = request.GET.get('fecha_inicio')
        fecha_final_str = request.GET.get('fecha_final')

        partos_qs = Parto.objects.all()
        rn_qs = RecienNacido.objects.all()
        pacientes_qs = Paciente.objects.all()

        if fecha_inicio_str and fecha_final_str:
            try:
                fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
                fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d').date()
                
                # CORRECCIÓN: Filtramos por la fecha CLÍNICA (fecha_hora), no la de sistema
                partos_qs = partos_qs.filter(fecha_hora__date__range=[fecha_inicio, fecha_final])
                rn_qs = rn_qs.filter(parto__fecha_hora__date__range=[fecha_inicio, fecha_final])
                
                ids = partos_qs.values_list('paciente_id', flat=True).distinct()
                pacientes_qs = Paciente.objects.filter(id__in=ids)
            except ValueError:
                pass
        else:
            ids = partos_qs.values_list('paciente_id', flat=True).distinct()
            pacientes_qs = Paciente.objects.filter(id__in=ids)

        return partos_qs, rn_qs, pacientes_qs, fecha_inicio_str, fecha_final_str


# --- VISTA DASHBOARD (PROTEGIDA) ---
class ReportesObstetriciaView(PermitsPositionMixin, ReporteFilterMixin, TemplateView):
    template_name = "reportes/dashboard_obstetricia.html"
    
    permission_required = ['READ_ONLY', 'ADMINISTRATIVE', 'CLINICAL_FULL', 'TOTAL_ACCESS']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        partos_qs, rn_qs, pacientes_qs, f_ini, f_fin = self.get_filtered_querysets(self.request)
        
        context.update({"fecha_inicio": f_ini, "fecha_final": f_fin})

        # Totales
        context["total_partos"] = partos_qs.count()
        context["total_recien_nacidos"] = rn_qs.count()
        context["total_complicaciones"] = partos_qs.exclude(Q(complicaciones__isnull=True) | Q(complicaciones__exact='')).count()

        # Listas Resumen
        # Agrupamos por el nombre del tipo de parto relacionado
        tipo_raw = partos_qs.values('tipo_parto__nombre').annotate(total=Count('id')).order_by('-total')
        context["distribucion_tipo_parto"] = [
            {'tipo_display': item['tipo_parto__nombre'], 'total': item['total']} 
            for item in tipo_raw if item['tipo_parto__nombre']
        ]

        # Posición
        pos_map = dict(Parto.PosicionPartoChoices.choices)
        pos_raw = partos_qs.values('posicion_parto').annotate(total=Count('id')).order_by('-total')
        context["distribucion_posicion_parto"] = [{'posicion_display': str(pos_map.get(i['posicion_parto'], i['posicion_parto'])), 'total': i['total']} for i in pos_raw if i['posicion_parto']]

        sex_map = dict(RecienNacido.SexoChoices.choices)
        sex_raw = rn_qs.values('sexo').annotate(total=Count('id')).order_by('sexo')
        context["distribucion_sexo_rn"] = [{'sexo_display': str(sex_map.get(i['sexo'], i['sexo'])), 'total': i['total']} for i in sex_raw]

        # Stats Vitales
        def calcular_stats(queryset, campo):
            vals = list(queryset.values_list(campo, flat=True).exclude(**{f'{campo}__isnull': True}))
            if not vals: return None, None
            prom = sum(vals) / len(vals)
            desv = sqrt(sum((x - prom) ** 2 for x in vals) / len(vals)) if len(vals) > 1 else 0
            return prom, desv

        p_peso, d_peso = calcular_stats(rn_qs, 'peso_gramos')
        p_talla, d_talla = calcular_stats(rn_qs, 'talla_cm')
        p_apgar, d_apgar = calcular_stats(rn_qs, 'apgar5')

        context.update({
            "promedio_peso_rn": p_peso, "stddev_peso_rn": d_peso,
            "promedio_talla_rn": p_talla, "stddev_talla_rn": d_talla,
            "promedio_apgar_rn": p_apgar, "stddev_apgar_rn": d_apgar,
        })

        # Demografía
        def get_dist_fk(qs, field):
            data = qs.values(f'{field}__nombre').annotate(total=Count('id')).order_by('-total')
            return [{'display': str(i[f'{field}__nombre']), 'total': i['total']} for i in data if i[f'{field}__nombre']]

        def get_dist_choice(qs, field, choices):
            mapping = dict(choices.choices)
            data = qs.values(field).annotate(total=Count('id')).order_by('-total')
            return [{'display': str(mapping.get(i[field], i[field])), 'total': i['total']} for i in data if i[field]]

        context["distribucion_pueblos"] = get_dist_fk(pacientes_qs, 'pueblo_originario')
        context["distribucion_nacionalidad"] = get_dist_fk(pacientes_qs, 'nacionalidad')
        context["distribucion_educacion"] = get_dist_choice(pacientes_qs, 'nivel_educacional', Paciente.NivelEducacionalChoices)
        context["distribucion_estado_civil"] = get_dist_choice(pacientes_qs, 'estado_civil', Paciente.EstadoCivilChoices)

        return context


# --- VISTA API (JSON) - PROTEGIDA ---
class ChartDataView(PermitsPositionMixin, View):
    permission_required = ['READ_ONLY', 'ADMINISTRATIVE', 'CLINICAL_FULL', 'TOTAL_ACCESS']

    def get(self, request):
        metric = request.GET.get('metric')
        days_param = request.GET.get('days', '7')
        start_date = None
        title_suffix = "(Histórico)"

        if days_param != 'historic':
            try:
                days = int(days_param)
                start_date = timezone.now().date() - timedelta(days=days)
                title_suffix = f"(Últimos {days} días)"
            except ValueError:
                pass

        data = {}

        # 1. EVOLUCIÓN (LÍNEA)
        if metric == 'partos_evolucion':
            qs = Parto.objects.all()
            # CORRECCIÓN: Usamos 'fecha_hora' (la real del parto)
            if start_date: qs = qs.filter(fecha_hora__date__gte=start_date)
            
            # CORRECCIÓN: Agrupamos por 'fecha_hora'
            evo = qs.annotate(date=TruncDate('fecha_hora')).values('date').annotate(c=Count('id')).order_by('date')
            chart = [[i['date'].strftime('%Y-%m-%d'), i['c']] for i in evo]
            data = {'title': f'Partos {title_suffix}', 'type': 'line', 'series_data': chart, 'labels': [x[0] for x in chart], 'values': [x[1] for x in chart]}

        elif metric == 'vitales_peso_evolucion':
            qs = RecienNacido.objects.all()
            # CORRECCIÓN: Usamos 'parto__fecha_hora' (la fecha del parto asociado)
            if start_date: qs = qs.filter(parto__fecha_hora__date__gte=start_date)
            
            # CORRECCIÓN: Agrupamos por 'parto__fecha_hora'
            evo = qs.annotate(date=TruncDate('parto__fecha_hora')).values('date').annotate(avg=Avg('peso_gramos')).order_by('date')
            chart = [[i['date'].strftime('%Y-%m-%d'), round(i['avg'], 1)] for i in evo]
            data = {'title': f'Peso Promedio {title_suffix}', 'type': 'line', 'color': '#34d399', 'series_data': chart, 'labels': [x[0] for x in chart], 'values': [x[1] for x in chart]}

        # 2. DISTRIBUCIONES (VARIEDAD)
        else:
            config_choice = {
                'complicaciones_distribucion': (Parto, 'complicaciones', None, 'doughnut'),
                'sexo_distribucion': (RecienNacido, 'sexo', RecienNacido.SexoChoices, 'pie'),
                'posicion_distribucion': (Parto, 'posicion_parto', Parto.PosicionPartoChoices, 'bar_horizontal'),
                'educacion_distribucion': (Paciente, 'nivel_educacional', Paciente.NivelEducacionalChoices, 'bar'),
                'estado_civil_distribucion': (Paciente, 'estado_civil', Paciente.EstadoCivilChoices, 'doughnut'),
            }
            
            config_fk = {
                'tipo_parto_distribucion': (Parto, 'tipo_parto', 'doughnut'),
                'pueblo_distribucion': (Paciente, 'pueblo_originario', 'pie'),
                'nacionalidad_distribucion': (Paciente, 'nacionalidad', 'bar_horizontal'),
            }

            if metric in config_fk:
                Model, field, chart_type = config_fk[metric]
                qs = Model.objects.all()
                
                # Filtro de fecha usando la fecha clínica correcta
                date_field = 'fecha_hora' if Model == Parto else ('parto__fecha_hora' if Model == RecienNacido else 'fecha_creacion')
                if Model == Paciente and start_date:
                    # Pacientes se filtran por sus partos en la fecha
                    partos_in = Parto.objects.filter(fecha_hora__date__gte=start_date)
                    qs = qs.filter(partos__in=partos_in).distinct()
                elif start_date:
                    # Nota: Si el modelo no tiene fecha_hora (como Paciente directo), usamos fecha_creacion por defecto,
                    # pero la lógica de arriba ya maneja Paciente via Parto.
                    qs = qs.filter(**{f'{date_field}__date__gte': start_date})

                # Agrupamos por nombre de la relación FK
                raw = qs.values(f'{field}__nombre').annotate(t=Count('id')).order_by('-t')
                chart_data = [{'name': i[f'{field}__nombre'], 'value': i['t']} for i in raw if i[f'{field}__nombre']]
                
                data = {'title': f'Distribución {title_suffix}', 'series_data': chart_data, 'type': chart_type}
                if 'bar' in chart_type:
                    data['labels'] = [x['name'] for x in chart_data]; data['values'] = [x['value'] for x in chart_data]

            elif metric in config_choice:
                Model, field, Choices, chart_type = config_choice[metric]
                qs = Model.objects.all()
                
                date_field = 'fecha_hora' if Model == Parto else ('parto__fecha_hora' if Model == RecienNacido else 'fecha_creacion')

                if Model == Paciente and start_date:
                    partos_in = Parto.objects.filter(fecha_hora__date__gte=start_date)
                    qs = qs.filter(partos__in=partos_in).distinct()
                elif start_date:
                    qs = qs.filter(**{f'{date_field}__date__gte': start_date})

                if field == 'complicaciones':
                     qs = qs.exclude(Q(complicaciones__isnull=True)|Q(complicaciones__exact=''))
                else:
                     qs = qs.exclude(**{f'{field}__exact': ''})

                raw = qs.values(field).annotate(t=Count('id')).order_by('-t')
                if field == 'complicaciones': raw = raw[:6]

                mapping = dict(Choices.choices) if Choices else {}
                chart_data = []
                for item in raw:
                    val = item[field]
                    name = str(mapping.get(val, val))
                    chart_data.append({'name': name, 'value': item['t']})
                
                data = {'title': f'Distribución {title_suffix}', 'series_data': chart_data, 'type': chart_type}
                if 'bar' in chart_type:
                    data['labels'] = [x['name'] for x in chart_data]; data['values'] = [x['value'] for x in chart_data]

        return JsonResponse(data)


# --- VISTA EXPORTAR EXCEL ---
class ExportarReporteExcelView(PermitsPositionMixin, ReporteFilterMixin, View):
    permission_required = ['READ_ONLY', 'ADMINISTRATIVE', 'CLINICAL_FULL', 'TOTAL_ACCESS']

    def get(self, request):
        partos_qs, rn_qs, _, f_ini, f_fin = self.get_filtered_querysets(request)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resumen Obstétrico"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        
        ws['A1'] = f"Reporte: {f_ini or 'Histórico'} - {f_fin or 'Hoy'}"
        ws['A3'] = "Total Partos"; ws['B3'] = partos_qs.count()
        
        cols = ["Fecha", "Paciente", "Edad", "Pueblo", "Educación", "Tipo", "Posición", "Sexo RN", "Peso", "Talla"]
        for idx, title in enumerate(cols, 1):
            c = ws.cell(row=6, column=idx, value=title)
            c.font = header_font; c.fill = header_fill

        row = 7
        for p in partos_qs.select_related('paciente', 'tipo_parto', 'paciente__pueblo_originario', 'paciente__nacionalidad').prefetch_related('recien_nacidos'):
            rn = p.recien_nacidos.first()
            edad = "N/A"
            if p.paciente.fecha_nacimiento:
                hoy = timezone.now().date()
                fn = p.paciente.fecha_nacimiento
                edad = hoy.year - fn.year - ((hoy.month, hoy.day) < (fn.month, fn.day))

            # Nombres seguros de FKs
            pueblo = p.paciente.pueblo_originario.nombre if p.paciente.pueblo_originario else "-"
            tipo_p = p.tipo_parto.nombre if p.tipo_parto else "-"

            ws.append([
                p.fecha_hora.strftime('%d/%m/%Y %H:%M'), # CORRECCIÓN: Usar fecha_hora
                str(p.paciente), edad,
                pueblo, p.paciente.get_nivel_educacional_display(),
                tipo_p, p.get_posicion_parto_display(),
                rn.get_sexo_display() if rn else "-", rn.peso_gramos if rn else "-", rn.talla_cm if rn else "-"
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Reporte_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response

# --- VISTA EXPORTAR PDF ---
class ExportarReportePDFView(PermitsPositionMixin, ReporteFilterMixin, View):
    permission_required = ['READ_ONLY', 'ADMINISTRATIVE', 'CLINICAL_FULL', 'TOTAL_ACCESS']
    
    def get(self, request):
        view = ReportesObstetriciaView()
        view.request = request
        context = view.get_context_data()
        template = get_template('reportes/reporte_pdf.html')
        html = template.render(context)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="Reporte.pdf"'
        pisa.CreatePDF(html, dest=response)
        return response
    

# --- VISTA AUDITORÍA (MODIFICADA CON BUSCADOR) ---
class ReporteAuditoriaView(PermitsPositionMixin, TemplateView):
    template_name = "reportes/auditoria_list.html"
    permission_required = ['ADMINISTRATIVE', 'TOTAL_ACCESS']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Filtros
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_final = self.request.GET.get('fecha_final')
        search_query = self.request.GET.get('q', '').strip() # <--- BUSCADOR
        
        qs = HistorialPaciente.objects.select_related('paciente', 'usuario').all()

        # 1. Filtro por Texto (Nuevo)
        if search_query:
            qs = qs.filter(
                Q(paciente__nombres__icontains=search_query) |
                Q(paciente__apellido_paterno__icontains=search_query) |
                Q(paciente__rut__icontains=search_query) |
                Q(usuario__username__icontains=search_query) |
                Q(usuario__first_name__icontains=search_query)
            )

        # 2. Filtro por Fecha
        if fecha_inicio and fecha_final:
            try:
                f_ini = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                f_fin = datetime.strptime(fecha_final, '%Y-%m-%d').date()
                qs = qs.filter(fecha__date__range=[f_ini, f_fin])
                context['fecha_inicio'] = fecha_inicio
                context['fecha_final'] = fecha_final
            except ValueError:
                pass
        
        context['historial_qs'] = qs[:200]
        context['q'] = search_query # <--- Para mantener el texto en el input
        return context